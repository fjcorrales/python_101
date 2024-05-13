# vamos a hacer un programa para procesar excels, algo simple, una modificacion de valores dentro de columnas
# y crear un grafico

import openpyxl as xl  # le ponemos un alias para poder hacer referencia a la libreria con el nombre que le pongamos
from openpyxl.chart import BarChart, Reference  # importamos las librerias necesarias para hacer un grafico


def procesamiento_excel(nombre):
    wb = xl.load_workbook(nombre)  # esta funcion abre el excel
    sheet = wb['Hoja 1']  # esto nos permite acceder a las paginas dentro del ecel

    # para acceder a una columna lo podemos hacer de dos formas
    # columna = sheet['a1']  #con [] y poniendo la coordenada exacta representada en el excel
    # sheet.cell(1, 1)    #con la funcion cell() y pasandole la coordenada generica de una matriz
    # print(sheet.max_row)    #nos devuelve el numero maximo de filas que tiene el excel

    for fila in range(2,
                      sheet.max_row + 1):  # el in range excluye el ultimo valor pasado como argumento, por eso el +1 y empezamos en 2 para ignorar los nombres de las columnas
        columna = sheet.cell(fila, 3)
        precio_corregido = columna.value * 0.9  # corregimos el valor del precio almacenado en la tercera columna ya que estos han de estar al 90%
        precio_ok = sheet.cell(fila, 4)
        precio_ok.value = precio_corregido

    valores = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4,
                        max_col=4)  # establecemos el rango de valores que queremos que use en el grafico
    grafico = BarChart()  # creamos el grafico
    grafico.add_data(valores)  # llenamos el grafico
    sheet.add_chart(grafico, 'e2')  # adjuntamos el grafico

    wb.save(nombre)  # con esto guardamos el excel


procesamiento_excel('transacciones.xlsx')
